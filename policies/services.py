import json
import logging
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date
from typing import Any, Dict, List

import openpyxl
from dateutil import parser
from dateutil.relativedelta import relativedelta
from django.db import transaction
from django.forms import ValidationError

from clients.enums import MaritalStatus
from config.models import BusinessSector
from core.enums import PremiumFrequency
from core.utils import get_dict_values, merge_dict_into_another, replace_keys, get_current_schema
from integrations.utils import calculate_binder_fees_amount, calculate_commission_amount, \
    calculate_guard_risk_admin_amount
from policies.constants import DEFAULT_CLIENT_FIELDS, DEFAULT_POLICY_FIELDS, POLICY_CLIENTS_COLUMNS_THF_UPDATE, \
    DEFAULT_BENEFICIARY_FIELDS
from policies.constants import FUNERAL_POLICY_CLIENT_COLUMNS
from policies.models import Policy
from policies.serializers import BeneficiarySerializer
from policies.serializers import ClientPolicyRequestSerializer, PremiumPaymentSerializer

logger = logging.getLogger(__name__)


def custom_serializer(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    elif obj == 'NULL':
        return None
    elif obj is None:
        return None


def serialize_for_jsonfield(data):
    for key, value in data.items():
        if isinstance(value, (datetime, date)):
            data[key] = value.isoformat()  # Convert datetime to ISO 8601 string
        elif value == 'NULL':
            data[key] = None  # Convert 'NULL' string to null
        elif value is None:
            data[key] = None
    return data


def extract_json_fields(dictionary):
    policy_details = {}

    for key in list(dictionary.keys()):
        if key.startswith("json_"):
            new_key = key.replace("json_", "")
            policy_details[new_key] = dictionary.pop(key)

    policy_details = serialize_for_jsonfield(policy_details)
    dictionary["policy_details"] = policy_details
    return json.dumps(dictionary, default=custom_serializer, indent=4)


@transaction.atomic
def upload_clients_and_policies(
        file_obj: Any, client_columns: List, policy_columns, source
) -> None:
    print("The columns loaded")
    # Convert received column definitions from JSON strings
    # received_policy_columns = json.loads(policy_columns)
    # received_client_columns = json.loads(client_columns)

    policy_type = None
    if source == "guardrisk":
        policy_type = 1
    elif source == "bordrex":
        policy_type = 1

    received_policy_columns = policy_columns
    received_client_columns = client_columns

    # Merge client and policy column definitions
    merged_columns = {**received_client_columns, **received_policy_columns}

    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_obj.file)

    # Extract expected column headers for clients and policies
    expected_client_headers: List[str] = get_dict_values(received_client_columns)
    expected_policy_headers: List[str] = get_dict_values(received_policy_columns)

    # Select the active worksheet
    ws = wb.active
    print(f'active sheet: {ws}')

    # Extract headers from the worksheet
    headers: List[str] = [cell.value.strip() for cell in ws[1]]

    # Check if expected headers are present in the worksheet
    expected_headers = expected_client_headers + expected_policy_headers
    if not set(expected_headers).issubset(set(headers)):
        unique_to_list1 = set(headers) - set(expected_headers)
        unique_to_list2 = set(expected_headers) - set(headers)
        print(f'unique to list1: {unique_to_list1}')
        print(f'unique to list2: {unique_to_list2}')

        raise ValidationError("Headers not matching the ones on the excel sheet")

    # Iterate over rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Create dictionary mapping headers to row values
        row_dict: Dict[str, Any] = dict(zip(headers, row))
        # Replace column keys with expected keys
        row_dict = replace_keys(merged_columns, row_dict)

        # Extract client and policy data

        client_data = {k: row_dict[k] for k in received_client_columns}
        client_data = merge_dict_into_another(client_data, DEFAULT_CLIENT_FIELDS)

        if "gender" in client_data:
            gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female", "Male": "Male", "Female": "Female"}
            client_data["gender"] = gender_mapping.get(client_data["gender"], "Unknown")

        policy_data = {k: row_dict[k] for k in received_policy_columns}

        policy_data = merge_dict_into_another(policy_data, DEFAULT_POLICY_FIELDS)
        # convert all fields that have json prefix
        policy_data = extract_json_fields(policy_data)
        if policy_type is not None:
            policy_data["policy_type"] = policy_type

        serializer = ClientPolicyRequestSerializer(
            data={"client": client_data, "policy": policy_data},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()


@transaction.atomic
def upload_funeral_clients_and_policies(
        file_obj: Any
) -> None:
    wb = openpyxl.load_workbook(file_obj.file)

    # with ThreadPoolExecutor() as executor:
    #     client_policy_future = executor.submit(process_worksheet, wb, "Funeral All", FUNERAL_POLICY_CLIENT_COLUMNS,
    #                                            "client_policy")
    #
    #     policy_beneficiary_future = executor.submit(process_worksheet, wb, "Beneficiary Details",
    #                                                 FUNERAL_POLICY_BENEFICIARY_COLUMNS,
    #                                                 "beneficiary")
    #
    #     client_policy_data = client_policy_future.result()
    #
    #     policy_beneficiary_data = policy_beneficiary_future.result()
    client_policy_data = process_worksheet(wb, "Funeral", FUNERAL_POLICY_CLIENT_COLUMNS, "client_policy")

    # policy_beneficiary_data = process_worksheet(wb, "Beneficiary Details", FUNERAL_POLICY_BENEFICIARY_COLUMNS,
    #                                             "beneficiary")

    # client_policy_beneficiary_data = match_beneficiaries_to_policies(policy_beneficiary_data, client_policy_data)
    #
    # client_policy_beneficiary_dependents_data = extract_funeral_dependant_fields(client_policy_beneficiary_data)

    # save_client_policy_beneficiary_dependents_data(client_policy_beneficiary_dependents_data)
    # save_client_policy_beneficiary_dependents_data(client_policy_data)
    policies = []
    for data in client_policy_data:
        print(f' data coming {data}')

        policy_number = data["policy_number"].strip() if isinstance(data["policy_number"],str) else data["policy_number"]
        print(f'policy')
        policy = Policy.objects.get(policy_number=policy_number, policy_type_id=2)
        if policy.policy_status in ['L', 'X', ]:
            date_obj = datetime.strptime(data['policy_details']['policy_status_date'], '%Y-%m-%dT%H:%M:%S')
            policy.closed_date = date_obj.strftime('%Y-%m-%d')
            policies.append(policy)
    Policy.objects.bulk_update(policies, ['closed_date'])
    # save_policy_beneficiary_data(policy_beneficiary_data)
    # client_columns = {
    #     "first_name": "principal_firstname",
    #     "last_name": "principal_surname",
    #     "primary_id_number": "principal_id",
    #     "gender": "principal_gender",
    #     "date_of_birth": "principal_date_of_birth",
    #     "email": "principal_email",
    #     "phone_number": "principal_phone_number",
    #     "address_street": "principal_physical_address",
    #     "postal_code": "principal_postal_code",
    #     "marital_status": "spouse_indicator",
    #     "primary_id_document_type": 1,
    #     "entity_type": "Individual"
    # }
    # dump_policies, dump_clients = extract_from_dump(client_policy_data, client_columns)
    # updated_clients = extract_employment_info_from_client(dump_clients)
    # save_client_policy_members_data(dump_policies, updated_clients)


def extract_funeral_json_fields(dictionary):
    details = {}
    for key in list(dictionary.keys()):
        if key.startswith("json_"):
            new_key = key.replace("json_", "")
            details[new_key] = dictionary.pop(key)

    if "policy_number" in dictionary:
        dictionary["policy_details"] = details
        return dictionary
    else:
        return dictionary


def extract_funeral_dependant_fields(client_details) -> List[Dict[str, Any]]:
    client_policy_beneficiary_dependents = []
    for client in client_details:
        dependents_list = []
        added_dependents = set()
        added_spouses = set()
        for key, value in client["policy_details"].items():
            if key.startswith("dependent") and value != "" and value is not None:
                # Check if key starts with "dependent", value is not an empty string, and not None
                dependent_number = key.split("_")[1]  # Extract the dependent number from the key
                if dependent_number not in added_dependents:
                    date_of_birth_key = f"dependent_{dependent_number}_date_of_birth "
                    date_of_birth = client["policy_details"][date_of_birth_key]
                    if isinstance(date_of_birth, int):
                        date_of_birth = "1900-01-01"

                    if date_of_birth is not None:  # Check if date_of_birth is not None
                        gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female"}
                        dependent = {
                            "number": dependent_number,
                            "dependant_gender": gender_mapping.get(
                                client["policy_details"].get(f"dependent_{dependent_number}_gender "), "Unknown"),
                            "dependant_dob": parser.parse(str(date_of_birth)).strftime(
                                "%Y-%m-%d") if date_of_birth != "1900-01-01" else None,
                            "relationship": 1
                        }
                        dependents_list.append(dependent)
                        added_dependents.add(dependent_number)
            if key.startswith("spouse") and value != "" and value is not None:
                date_of_birth_key = f"spouse_firstname "
                spouse_name = client["policy_details"][date_of_birth_key]
                spouse_dob = client["policy_details"].get(f"spouse_date_of_birth ")
                if isinstance(spouse_dob, int):
                    spouse_dob = "1900-01-01"
                if spouse_name is not None and spouse_name not in added_spouses:  # Check if date_of_birth is not None
                    gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female"}
                    dependent = {
                        "dependant_name": str(client["policy_details"].get(f"spouse_firstname ")) + " " + str(
                            client["policy_details"].get(f"spouse_surname ")),
                        "dependant_gender": gender_mapping.get(client["policy_details"].get(f"spouse_gender "),
                                                               "Unknown"),
                        "dependant_dob": parser.parse(str(spouse_dob)).strftime(
                            "%Y-%m-%d") if spouse_dob != "1900-01-01" else None,
                        "relationship": 2
                    }
                    dependents_list.append(dependent)
                    added_spouses.add(spouse_name)

        # Print the list of dependents

        client["dependants"] = dependents_list
        client_policy_beneficiary_dependents.append(client)
    return client_policy_beneficiary_dependents


def process_worksheet(
        wb: openpyxl.Workbook, worksheet_name: str, columns: List[str], data_type: str
) -> List[Dict[str, Any]]:
    worksheet = wb[worksheet_name]
    print(f'worksheet name: {worksheet}')
    if worksheet.max_row < 1:
        raise ValidationError(f"The worksheet {worksheet_name} is empty or has no header row")
    headers = [cell.value.strip() for cell in worksheet[1] if cell.value]
    expected_headers = get_dict_values(columns)
    if not set(expected_headers).issubset(set(headers)):
        unique_to_actual_headers = set(headers) - set(expected_headers)
        unique_to_expected_headers = set(expected_headers) - set(headers)
        print(f'unique to expected headers: {unique_to_expected_headers}')
        print(f'unique actual headers: {unique_to_actual_headers}')
        raise ValidationError(f"{data_type.capitalize()} headers not matching the ones on the excel sheet")

    processed_data = []
    # for row in worksheet.iter_rows(min_row=2, values_only=True):
    #     if any(cell is not None for cell in row):
    #         row_dict = dict(zip(headers, row))
    #         if row_dict is not None:
    #             row_dict = replace_keys(columns, row_dict)
    #             if row_dict is not None:
    #                 data = {k: row_dict[k] for k in columns if k in row_dict}
    #                 if data is not None:
    #                     print(f'data {data}')
    #                     policy_number = data["loanId"]
    #                     print(f'policy number: {policy_number}')
    #                     policy = Policy.objects.filter(policy_number=policy_number).first()
    #                     if policy is not None:
    #                         if policy.premium is None:
    #                             print(f'policy {policy}')
    #                             premium = policy.total_premium
    #                             total_premium = data["total_premium"]
    #                             policy.premium = premium
    #                             policy.total_premium = total_premium
    #                             policy.save()
    # for row in worksheet.iter_rows(min_row=2, values_only=True):
    #     if any(cell is not None for cell in row):
    #         row_dict = dict(zip(headers, row))
    #         if row_dict is not None:
    #             row_dict = replace_keys(columns, row_dict)
    #             if row_dict is not None:
    #                 data = {k: row_dict[k] for k in columns if k in row_dict}
    #                 if data is not None:
    #                     print(f'data {data}')
    #                     #policy_number, _ = get_policy_number_and_external_id(data)
    #                     policy_number= data['policy_number']
    #                     print(f'policy number: {policy_number}')
    #                     policy = Policy.objects.filter(policy_number=policy_number).first()
    #                     if policy is not None:
    #                         print(f'policy {policy}')
    #                         policy.policy_status = map_closure_reason(data["closed_reason"])
    #                         policy.closed_date = data["expiry_date"]
    #                         policy.save()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            row_dict = dict(zip(headers, row))
            row_dict = replace_keys(columns, row_dict)
            data = {k: row_dict[k] for k in columns if k in row_dict}
            if data_type == "client_policy":
                default_columns = {**DEFAULT_CLIENT_FIELDS, **DEFAULT_POLICY_FIELDS,
                                   "entity": "Indlu",
                                   "commission_frequency": "Monthly",
                                   "premium_frequency": "Monthly",
                                   "policy_type_id": 2,
                                   "commission_percentage": 7.50
                                   }
            elif data_type == "policy":
                default_columns = {**DEFAULT_POLICY_FIELDS, "entity": "Indlu",
                                   "sub_scheme": "Credit Life",
                                   "commission_frequency": "Monthly",
                                   "premium_frequency": "Monthly",
                                   }
            elif data_type == "client":
                default_columns = {**DEFAULT_CLIENT_FIELDS}
            elif data_type == "repayment":
                default_columns = {}
            elif data_type == "policy_client_dump" or data_type == 'cfsa' or data_type == 'cfsacorrect':
                default_columns = {**DEFAULT_POLICY_FIELDS, **DEFAULT_CLIENT_FIELDS,
                                   "entity": "Indlu",
                                   "policy_type_id": 1,
                                   "sub_scheme": "Credit Life",
                                   "commission_frequency": "Monthly",
                                   "premium_frequency": "Monthly",
                                   "commission_percentage": 7.50
                                   }
            else:
                default_columns = DEFAULT_BENEFICIARY_FIELDS

            data = merge_dict_into_another(data, default_columns)
            processed_data.append(process_data(data, data_type))
    return processed_data


def process_data(data: Dict[str, Any], data_type: str) -> Dict[str, Any]:
    processed_data = None
    if data_type == "client_policy":
        processed_data = process_client_policy_data(data)
    elif data_type == "policy":
        processed_data = process_indlu_policy_data(data)
    elif data_type == "client":
        processed_data = process_indlu_client_data(data)
    elif data_type == "policy_client_dump" or data_type == 'cfsa' or data_type == 'cfsacorrect':
        processed_data = process_indlu_policy_client_dump_data(data)
    elif data_type == "beneficiary":
        processed_data = process_beneficiary_data(data)
    elif data_type == "repayment":
        processed_data = process_indlu_repayment_data(data)
    else:
        logger.warning(f"Unknown data type: {data_type}")
    return processed_data


def process_indlu_client_data(client_data: Dict[str, Any]) -> Dict[str, Any]:
    client = extract_employment_fields(client_data)
    gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female", "Male": "Male", "Female": "Female"}
    client["gender"] = gender_mapping.get(client.get("gender"), "Unknown")
    return client


def extract_employment_fields(client_details):
    employment_fields = {
        'job_title',
        'sector',
        'employer',
        'employment_date',
        'employer_name',
        'gross_salary',
        'basic_salary',
    }
    employment_data = {}
    for field in employment_fields:
        if field in client_details:
            employment_data[field] = client_details.pop(field)
            client_details['employment_details'] = employment_data
    return client_details


def process_indlu_policy_data(policy_data: Dict[str, Any]):
    # Further processing if needed
    policy_data["policy_term"] = 1 if policy_data["policy_term"] == 0 else policy_data[
        "policy_term"]
    __calculate_and_set_expiry_date(policy_data)
    extract_json_fields(policy_data)
    return policy_data


def process_indlu_policy_client_dump_data(policy_data: Dict[str, Any]):
    print(f'policy data {policy_data}')
    if 'business_unit' not in policy_data:
        policy_data['business_unit'] = 'THF'
        policy_data["product_name"] = 'Housing Loans (Take on)'
        policy_data["sub_scheme"] = 'Credit Life'

    if 'policy_status' in policy_data:
        status = policy_data['policy_status']
        if status == 'F':
            policy_data['policy_status'] = 'P'
    if 'address_street' in policy_data:
        address_street = policy_data['address_street']
        if isinstance(address_street, datetime):
            policy_data['address_street'] = address_street.strftime('%d/%m/%Y')
    if 'date_of_birth' in policy_data:
        date_of_birth = policy_data['date_of_birth']
        if isinstance(date_of_birth, str):
            if date_of_birth == 'NULL':
                policy_data['expiry_date'] = None
            else:
                try:
                    policy_data['date_of_birth'] = parser.parse(date_of_birth).date()
                except ValueError:
                    policy_data['date_of_birth'] = date.today().strftime('%d/%m/%Y')
        if date_of_birth is None:
            policy_data['date_of_birth'] = date.today().strftime('%d/%m/%Y')
    if 'primary_id_number' in policy_data:
        primary_id_number = policy_data['primary_id_number']
        if primary_id_number is None:
            policy_data['primary_id_number'] = (policy_data['first_name'] + ' ' + policy_data['last_name']
                                                + str(policy_data['policy_number']))
    if 'commencement_date' in policy_data:
        commencement_date = policy_data['commencement_date']
        if isinstance(commencement_date, str):
            if commencement_date == 'NULL':
                policy_data['commencement_date'] = None
            else:
                policy_data['commencement_date'] = parser.parse(commencement_date)
    if 'expiry_date' in policy_data:
        date_of_birth = policy_data['expiry_date']
        if isinstance(date_of_birth, str):
            if date_of_birth == 'NULL':
                policy_data['expiry_date'] = None
            else:
                policy_data['expiry_date'] = parser.parse(date_of_birth)
    if 'address_province' in policy_data:
        address_street = policy_data['address_province']
        if address_street == 'NULL':
            policy_data['address_province'] = None
    if 'json_employer_payment_calender' in policy_data:
        calender = policy_data['json_employer_payment_calender']
        policy_data['json_employer_payment_calender'] = calender.replace("\\", "")
    if 'json_current_loan_balance' in policy_data:
        json_current_loan_balance = policy_data['json_current_loan_balance']
        if json_current_loan_balance == 'NULL' or json_current_loan_balance == '' or json_current_loan_balance is None:
            policy_data['json_current_loan_balance'] = 0
            policy_data['policy_status'] = "F"
        else:
            policy_data['policy_status'] = "A"
    if 'sum_insured' in policy_data:
        sum_insured = policy_data['sum_insured']
        if isinstance(sum_insured, str):
            policy_data['sum_insured'] = round(float(sum_insured.replace(',', '')), 2)

    __calculate_and_set_expiry_date(policy_data)
    extract_json_fields(policy_data)

    total_premium = policy_data.get("total_premium", 0)
    total_premium = 0 if total_premium == 'NULL' else total_premium
    policy_data["commission_amount"] = calculate_commission_amount(total_premium)
    policy_data["admin_fee"] = calculate_guard_risk_admin_amount(total_premium)
    policy_data["policy_details"]["binder_fees"] = calculate_binder_fees_amount(total_premium)
    return policy_data


def __calculate_and_set_expiry_date(policy_data: Dict[str, Any]):
    if policy_data["expiry_date"] == 'NULL' or policy_data["expiry_date"] == '':
        policy_term = int(policy_data["policy_term"])
        policy_data["expiry_date"] = policy_data["commencement_date"].date() + relativedelta(months=policy_term)
    elif isinstance(policy_data["expiry_date"], datetime):
        policy_data["expiry_date"] = policy_data["expiry_date"].date()


def process_indlu_repayment_data(payment_data: Dict[str, Any]) -> Dict[str, Any]:
    if 'payment_date' in payment_data:
        payment_date = payment_data['payment_date']
        if isinstance(payment_date, datetime):
            # If payment_date is already a datetime object, format it as needed
            payment_data['payment_date'] = payment_date.strftime('%Y-%m-%d')  # Reformat date if needed
        elif isinstance(payment_date, str):
            payment_date = parser.parse(payment_date)
            payment_data['payment_date'] = payment_date.date()
    if 'is_reversed' in payment_data:
        value = payment_data['is_reversed']
        # Convert Excel-style boolean strings to Python booleans
        if isinstance(value, str) and value.upper() == '=FALSE()':
            payment_data['is_reversed'] = False
        elif isinstance(value, str) and value.upper() == '=TRUE()':
            payment_data['is_reversed'] = True
    if 'amount' in payment_data:
        payment_date = payment_data['amount']
        if isinstance(payment_date, str):
            # If payment_date is already a datetime object, format it as needed
            payment_data['amount'] = round(float(payment_date), 2)  # Reformat date if needed
        elif isinstance(payment_date, float):
            payment_date = round(payment_date, 2)
            payment_data['amount'] = payment_date
    return payment_data


def process_beneficiary_data(beneficiary_data: Dict[str, Any]) -> Dict[str, Any]:
    # Further processing if needed

    beneficiary_data["beneficiary_name"] = str(beneficiary_data["beneficiary_first_name"]) + " " + str(
        beneficiary_data["beneficiary_last_name"])
    beneficiary_data["relationship"] = 1
    beneficiary_data.pop("beneficiary_first_name")
    beneficiary_data.pop("beneficiary_last_name")
    # policy_number = beneficiary_data["beneficiary_policy_number"]
    # try:
    #     policy = Policy.objects.filter(
    #         policy_number=policy_number
    #     )
    #     if not policy.exists():
    #         raise Exception(f"Policy {policy_number} does not exist")
    #
    #     beneficiary_data["policy_id"] = policy.get().id
    #
    # except Exception as e:
    #     logger.error(f"Policy with policy number {policy_number} not found: {e}")

    for detail in beneficiary_data:
        if isinstance(beneficiary_data[detail], datetime):
            beneficiary_data[detail] = beneficiary_data[detail].strftime('%Y-%m-%d')

    return beneficiary_data


def process_client_policy_data(client_policy_data: Dict[str, Any]) -> Dict[str, Any]:
    print(f'funeral policy data {client_policy_data}')
    if 'business_unit' not in client_policy_data:
        client_policy_data['business_unit'] = 'NiftyCover'
    if "gender" in client_policy_data:
        gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female", "Male": "Male", "Female": "Female"}
        client_policy_data["gender"] = gender_mapping.get(client_policy_data["gender"], "Unknown")
    frequency_map = {
        1: PremiumFrequency.MONTHLY,
        3: PremiumFrequency.QUARTERLY,
        12: PremiumFrequency.ANNUAL,
        6: PremiumFrequency.SEMI_ANNUAL,
        2: PremiumFrequency.BI_ANNUAL,
        0: PremiumFrequency.ONCE_OFF
    }

    marital_status_map = {
        "N": MaritalStatus.SINGLE,
        "Y": MaritalStatus.MARRIED
    }
    client_policy_data["premium_frequency"] = frequency_map.get(client_policy_data["premium_frequency"],
                                                                PremiumFrequency.ONCE_OFF).value

    client_policy_data["commission_frequency"] = frequency_map.get(client_policy_data["commission_frequency"],
                                                                   PremiumFrequency.ONCE_OFF).value
    client_policy_data["marital_status"] = marital_status_map.get(client_policy_data["marital_status"],
                                                                  MaritalStatus.SINGLE).value

    gender_mapping = {"U": "Unknown", "M": "Male", "F": "Female"}

    client_policy_data["gender"] = gender_mapping.get(client_policy_data.get("gender"), "Unknown")
    if 'policy_status' in client_policy_data:
        status = clean_string(client_policy_data['policy_status'])
        client_policy_data['policy_status'] = status
        if status in ['F']:
            client_policy_data['policy_status'] = 'P'
        if status in ['P,', 'L', 'X']:
            status_date = client_policy_data['json_policy_status_date']
            print(f'status date {status_date}')
            client_policy_data['closed_date'] = status_date
    if 'date_of_birth' in client_policy_data:
        expiry_date = client_policy_data['date_of_birth']
        if isinstance(expiry_date, str):
            if expiry_date == 'NULL':
                client_policy_data['date_of_birth'] = None
            else:
                try:
                    client_policy_data['date_of_birth'] = parser.parse(expiry_date).date()
                except ValueError:
                    client_policy_data['date_of_birth'] = date.today().strftime('%d/%m/%Y')
        if expiry_date is None:
            client_policy_data['date_of_birth'] = date.today().strftime('%d/%m/%Y')
    if 'commencement_date' in client_policy_data:
        commencement_date = client_policy_data['commencement_date']
        if isinstance(commencement_date, str):
            if commencement_date == 'NULL':
                client_policy_data['commencement_date'] = None
            else:
                client_policy_data['commencement_date'] = parser.parse(commencement_date)
    if 'expiry_date' in client_policy_data:
        expiry_date = client_policy_data['expiry_date']
        if isinstance(expiry_date, str):
            if expiry_date == 'NULL':
                client_policy_data['expiry_date'] = None
            else:
                client_policy_data['expiry_date'] = parser.parse(expiry_date)
    if 'address_street' in client_policy_data:
        address = client_policy_data['address_street']
        if address is not None:
            address_parts = address.split(';')
            street = address_parts[0].strip()
            city = address_parts[2].strip()
            client_policy_data['address_street'] = street
            client_policy_data['address_town'] = city
    if 'sum_insured' in client_policy_data:
        sum_insured = client_policy_data['sum_insured']
        if isinstance(sum_insured, str):
            client_policy_data['sum_insured'] = round(float(sum_insured.replace(',', '')), 2)

    __calculate_and_set_expiry_date(client_policy_data)
    extract_json_fields(client_policy_data)

    total_premium = client_policy_data.get("total_premium", 0)
    total_premium = 0 if total_premium == 'NULL' else total_premium
    client_policy_data["commission_amount"] = calculate_commission_amount(total_premium)
    client_policy_data["admin_fee"] = calculate_guard_risk_admin_amount(total_premium)
    client_policy_data["policy_details"]["binder_fees"] = calculate_binder_fees_amount(total_premium)
    return client_policy_data


def clean_string(input_str):
    # Replace all occurrences of [, ], and ' in one go
    return re.sub(r"[\[\]']", '', input_str).strip()


def save_policy_beneficiary(policy_beneficiary_data: Dict[str, Any]):
    # Filter out non-dictionary values and None values
    data = {
        key: value
        for key, value in policy_beneficiary_data.items()
        if isinstance(value, dict) and "policy_id" in value
    }

    serializer = BeneficiarySerializer(data=policy_beneficiary_data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    logger.info(f"Saved Policy Beneficiary {policy_beneficiary_data}")


def match_members_to_policies(members, policies):
    matched_policies = []
    for policy in policies:
        member_client_reference = policies['client_id']
        for member in members:
            if member['client_id'] == member_client_reference:
                policy['client'] = [member]
                matched_policies.append(policy)
    return matched_policies


@transaction.atomic
def save_client_policy_beneficiary_dependents_data(client_policy_data: List[Dict[str, Any]]) -> None:
    for client in client_policy_data:
        save_client_policy(client)
    # with ThreadPoolExecutor(max_workers=20) as executor:
    #     futures = [executor.submit(save_client_policy, client) for client in client_policy_data]
    #
    #     # Wait for all futures to complete
    #     for future in futures:
    #         future.result()


@transaction.atomic
def save_policy_beneficiary_data(policy_beneficiary_data: List[Dict[str, Any]]) -> None:
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = [executor.submit(save_policy_beneficiary, policy) for policy in policy_beneficiary_data]

        # Wait for all futures to complete
        for future in futures:
            future.result()


@transaction.atomic
def upload_bulk_repayments(
        file_obj: Any, repayment_columns
) -> None:
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_obj.file)
    repayments = process_worksheet(wb, "Receipts", repayment_columns, "repayment")
    save_indlu_repayments_data(repayments)


@transaction.atomic
def save_indlu_repayments_data(repayments: List[Dict[str, Any]]) -> None:
    failed_repayments = []
    for repayment in repayments:
        try:
            if not repayment["policy_id"]:
                print(f"Policy ID missing from request! : {repayment}")
                failed_repayments.append(repayment['policy_id'])
            else:
                serializer = PremiumPaymentSerializer(
                    data=repayment
                )
                serializer.is_valid(raise_exception=True)
                serializer.save()
        except Exception as e:
            print(f"Error saving {repayment['policy_id']}")
            print(e)
            failed_repayments.append(repayment['policy_id'])


@transaction.atomic
def upload_indlu_clients_and_policies(
        file_obj: Any, source
) -> None:
    schema = get_current_schema()
    print(f'schema: {schema}')
    wb = openpyxl.load_workbook(file_obj.file)
    # clients = process_worksheet(wb, "Members", CLIENT_COLUMNS_INDLU, "client")
    data_type = source
    if data_type == 'cfsa_update':
        policy_clients_dump = process_worksheet(wb, "DataDump", POLICY_CLIENTS_COLUMNS_THF_UPDATE, "cfsa_update")
    # elif data_type == 'cfsa':
    #     policy_clients_dump = process_worksheet(wb, "Data", POLICY_CLIENTS_COLUMNS_CFSA, "cfsa")
    # else:
    #     policy_clients_dump = process_worksheet(wb, "DataDump", POLICY_CLIENT_COLUMNS_INDLU, "policy_client_dump")
    # client_columns = {
    #     **CLIENT_COLUMNS_INDLU,
    #     "primary_id_document_type": 1,
    #     "entity_type": "Individual"}
    # dump_policies, dump_clients = extract_from_dump(policy_clients_dump, client_columns)
    # updated_clients = extract_employment_info_from_client(dump_clients)
    # # extract_and_save_sectors(dump_clients)
    # save_client_policy_members_data(dump_policies, updated_clients)
    print('Done saving policies and clients from dump')


def extract_and_save_sectors(dump_clients):
    for client in dump_clients:
        if 'employment_details' in client:
            sector = client["employment_details"]["sector"]
            db_sector, created = BusinessSector.objects.get_or_create(sector=sector)


def extract_employment_info_from_client(clients):
    updated_clients = []
    for client in clients:
        updated_client = process_indlu_client_data(client)
        updated_clients.append(updated_client)
    return updated_clients


def extract_from_dump(policy_clients_dump, received_client_columns):
    policies = []
    clients = []

    for row in policy_clients_dump:
        # Extract client information
        client = {col: row[col] for col in received_client_columns if col in row}
        clients.append(client)
        # Extract policy information
        policy = {col: row[col] for col in row if col not in received_client_columns or col == 'client_id'}
        policies.append(policy)

    return policies, clients


def match_clients_to_policies(policies, clients):
    matched_policies = []
    for policy in policies:
        client_id = policy['client_id']
        for client in clients:
            if policy['client_id'] == client_id:
                if policy.get('client') is None:
                    policy['client'] = client
                    matched_policies.append(policy)
    return matched_policies


def match_policies_and_clients(policies, clients, dump_policies, dump_clients):
    # Convert policies and clients list to dictionaries for quick lookup
    policy_dict = {policy["policy_number"]: policy for policy in policies}
    client_dict = {client["client_id"]: client for client in clients}

    updated_dump_policies = []
    updated_dump_clients = []

    # Helper function to merge dictionaries with preference for non-null values
    # def merge_with_preference(primary, secondary):
    #     return {key: primary.get(key) if primary.get(key) is not None else secondary.get(key) for key in
    #             set(primary) | set(secondary)}

    # Match and update dump_policies with extra fields from policies
    for dump_policy in dump_policies:
        policy_number = dump_policy["policy_number"]
        if policy_number in policy_dict:
            # Update dump_policy with extra fields from matching policy
            matched_policy = {**dump_policy, **policy_dict[policy_number]}
            updated_dump_policies.append(matched_policy)
            # Remove the matched policy from the policies list
            policies = [p for p in policies if p["policy_number"] != policy_number]
        else:
            updated_dump_policies.append(dump_policy)

    # Match and update dump_clients with extra fields from clients
    for dump_client in dump_clients:
        client_id = dump_client["client_id"]
        if client_id in client_dict:
            # Update dump_client with extra fields from matching client
            matched_client = {**dump_client, **client_dict[client_id]}
            updated_dump_clients.append(matched_client)
            # Remove the matched client from the clients list
            clients = [c for c in clients if c["client_id"] != client_id]
        else:
            updated_dump_clients.append(dump_client)

    # Combine updated dump_policies with remaining unmatched policies
    final_policies = updated_dump_policies + policies

    # Combine updated dump_clients with remaining unmatched clients
    final_clients = updated_dump_clients + clients

    return final_policies, final_clients


@transaction.atomic
def save_client_policy_members_data(policies, clients):
    for policy in policies:
        policy_exists = Policy.objects.filter(policy_number=policy["policy_number"]).exists()
        if not policy_exists:
            for client in clients:
                if client['client_id'] == policy['client_id']:
                    serializer = ClientPolicyRequestSerializer(data={"client": client, "policy": policy})
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                    print(f"Saved Client and Policy")
                    break


@transaction.atomic
def save_client_policy_members_receipts_cli_charges_data(client_policy_data: List[Dict[str, Any]]) -> None:
    for client in client_policy_data:
        save_client_policy(client)


def save_client_policy(client_policy_data):
    print(f'funeral data {client_policy_data}')
    # client = {key: client_policy_data.pop(key) for key in ["first_name",
    #                                                        "last_name",
    #                                                        "primary_id_number",
    #                                                        "gender",
    #                                                        "date_of_birth",
    #                                                        "email",
    #                                                        "phone_number",
    #                                                        "address_street",
    #                                                        "address_town",
    #                                                        "postal_code",
    #                                                        "marital_status",
    #                                                        "primary_id_document_type",
    #                                                        "entity_type"]}
    # policy_details = client_policy_data["policy_details"]
    # for key, value in policy_details.items():
    #     if isinstance(value, time):
    #         policy_details[key] = value.strftime("%H:%M:%S")
    # # Serialize to JSON
    # client_policy_data["policy_details"] = json.dumps(policy_details)
    #
    # serializer = ClientPolicyRequestSerializer(
    #     data={"client": client, "policy": client_policy_data},
    # )
    # serializer.is_valid(raise_exception=True)
    # serializer.save()
    logger.info(f"Saved Client and Policy {client} {client_policy_data}")


def match_beneficiaries_to_policies(beneficiaries, policies):
    matched_policies = []
    for beneficiary in beneficiaries:
        beneficiary_policy_number = beneficiary['beneficiary_policy_number']

        for policy in policies:
            if policy['policy_number'] == beneficiary_policy_number:
                if policy.get('beneficiaries') is None:
                    policy['beneficiaries'] = [beneficiary]
                else:
                    policy['beneficiaries'].append(beneficiary)
                matched_policies.append(policy)

    return matched_policies
