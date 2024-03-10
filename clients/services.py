import json
import logging
from django.db import transaction
from django.forms import ValidationError
import openpyxl
from typing import Any, Dict, List
from clients.models import ClientDetails
from rest_framework.serializers import ValidationError as DRFValidationError
from clients.serializers import ClientDetailsSerializer
from core.utils import get_dict_values, replace_keys

logger = logging.getLogger(__name__)


def upload_clients(file_obj: Any, columns: List) -> None:

    received_columns = json.loads(columns)
    wb = openpyxl.load_workbook(file_obj.file)

    # get the values that match the columns of the excel
    expected_headers: List[str] = get_dict_values(received_columns)

    # Select the first worksheet
    ws = wb.active

    # Get the headers from the worksheet
    headers: List[str] = [cell.value.strip() for cell in ws[1]]

    # Validate the headers
    if headers != expected_headers:
        raise ValidationError("Headers not matching the ones on the excel sheet")

    # Create an empty list to store validated data
    validated_data_list: List[Dict[str, Any]] = []

    # Loop through the rows in the worksheet and validate each row using the serializer
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: Dict[str, Any] = dict(zip(headers, row))
        # switch columns with the expected
        row_dict = replace_keys(received_columns, row_dict)
        print(row_dict)
        serializer = ClientDetailsSerializer(data=row_dict)
        try:
            serializer.is_valid(raise_exception=True)
            print("After is_valid")
            validated_data_list.append(serializer.validated_data)
        except DRFValidationError as e:
            print("drf error")
            raise ValidationError(e)

    # Bulk create the objects within a transaction
    with transaction.atomic():
        ClientDetails.objects.bulk_create(
            [ClientDetails(**validated_data) for validated_data in validated_data_list]
        )
