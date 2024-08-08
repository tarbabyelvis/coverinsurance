from collections import defaultdict
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Font

from integrations.guardrisk.data.premiums import calculate_binder_fee_amount, calculate_insurer_commission_amount
from integrations.utils import is_new_policy, calculate_nett_amount, calculate_vat_amount, \
    calculate_guard_risk_admin_amount, calculate_amount_excluding_vat


def bordrex_report_util(policies, from_date, to_date, entity):
    timestamp = date.today().strftime("%Y%m%d")
    return populate_policies(policies, from_date, to_date, entity, timestamp)


def generate_main_report(policies):
    aggregates = defaultdict(lambda: [0, 0, 0, 0, 0])
    for policy in policies:
        division = policy['division']
        try:
            risk_premium = float(policy['premium'])
        except TypeError as e:
            risk_premium = 0
        else:
            try:
                guardrisk_fee = float(policy.get("net_premium_paid_to_gr", 0))
            except ValueError as e:
                guardrisk_fee = calculate_guard_risk_admin_amount(premium_amount=risk_premium)
            try:
                commission = float(policy.get("commission_amount", 0))
            except ValueError as e:
                commission = calculate_insurer_commission_amount(premium_amount=risk_premium)
            try:
                binder_fee = float(policy.get("binder_fee", 0))
            except ValueError as e:
                binder_fee = calculate_binder_fee_amount(premium_amount=risk_premium)
            try:
                nett = float(policy.get("net_premium_paid_to_gr", 0))
            except ValueError as e:
                nett = calculate_nett_amount(
                    premium_amount=risk_premium,
                    guardrisk_amount=guardrisk_fee,
                    commission=commission,
                    binder_fee=binder_fee)

            aggregates[division][0] += risk_premium
            aggregates[division][1] += guardrisk_fee
            aggregates[division][2] += commission
            aggregates[division][3] += binder_fee
            aggregates[division][4] += nett
    sub_totals = {
        "division": "",
        "total_risk_premium": 0,
        "total_guardrisk_fee": 0,
        "total_commission": 0,
        "total_binder_fee": 0,
        "total_nett": 0
    }
    for division, totals in aggregates.items():
        sub_totals["division"] = division
        sub_totals["total_risk_premium"] += totals[0]
        sub_totals["total_guardrisk_fee"] += totals[1]
        sub_totals["total_commission"] += totals[2]
        sub_totals["total_binder_fee"] += totals[3]
        sub_totals["total_nett"] += totals[4]

    sub_totals_sum = [
        "",
        "{:,.2f}".format(sub_totals["total_risk_premium"]),
        "{:,.2f}".format(sub_totals["total_guardrisk_fee"]),
        "{:,.2f}".format(sub_totals["total_commission"]),
        "{:,.2f}".format(sub_totals["total_binder_fee"]),
        "{:,.2f}".format(sub_totals["total_nett"])
    ]
    vat_total_binder_fee = calculate_vat_amount(sub_totals["total_binder_fee"])
    vat_total_nett = calculate_vat_amount(sub_totals["total_nett"])
    vat_totals_sum = [
        "VAT Amount included @15%",
        "-",
        "-",
        "-",
        "{:,.2f}".format(vat_total_binder_fee),
        "{:,.2f}".format(vat_total_nett)
    ]
    total_binder_fee: float = sub_totals["total_binder_fee"]
    total_nett_sub: float = sub_totals["total_nett"]
    totals_sum = [
        "",
        "{:,.2f}".format(sub_totals["total_risk_premium"]),
        "{:,.2f}".format(sub_totals["total_guardrisk_fee"]),
        "{:,.2f}".format(sub_totals["total_commission"]),
        "{:,.2f}".format(total_binder_fee),
        "{:,.2f}".format(total_nett_sub)
    ]
    totals = [sub_totals_sum, vat_totals_sum, totals_sum]
    return aggregates, totals


def calculate_total(*amounts):
    return sum(amounts)


def generate_excel_report_util(policy_payments, from_date, to_date, entity):
    wb = openpyxl.Workbook()
    ws_front_sheet = wb.active
    ws_front_sheet.title = "Borderaux Report"
    ws_policies_sheet = wb.create_sheet(title="Data")
    timestamp = date.today().strftime("%Y%m%d")
    client_identifier = get_client_identifier(entity)
    policy_payments = populate_policies(policy_payments, from_date, to_date, entity, timestamp)
    reports, totals = generate_main_report(policy_payments)
    generate_template_data(reports, totals, from_date, to_date, ws_front_sheet, f'L00{client_identifier}')
    header = ["Time Stamp", "Period Start", "Period End", "Administrator Identifier", "Insurer Name",
              "Client Identifier", "Division", "Risk Identifier", "Sub Scheme Name", "Policy Number",
              "Commencement Date", "Expiry Date", "New Business Indicator", "Policy Term", "Total Premium",
              "Premium Paid", "Commission", "Binder Fee", "Net Premium Paid Over To GR Including GR Admin Fee",
              "Admin Fee", "Premium Frequency", "Death Indicator", "PTD Indicator", "TTD Indicator",
              "Retrenchment Indicator", "Dread Disease Indicator", "Identity Theft Indicator",
              "Accidental Death Indicator", "Original Loan Balance", "Current Outstanding Balance",
              "Installment Amount",
              "Life 1 Last Name", "Life 1 First Name", "Life 1 Initials", "Life 1 ID", "Life 1 Gender",
              "Life 1 Date Of Birth", "Life 1 Date Of Death", "Life 1 Physical Address",
              "Life 1 City", "Life 1 Province", "Postal Code", "Life 1 Phone Number"
              ]
    ws_policies_sheet.append(header)
    policies_data = convert_dictionary_to_list(policy_payments)
    for row in policies_data:
        ws_policies_sheet.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Rewind the buffer to the beginning
    return buffer.getvalue()


def generate_template_data(reports, totals, from_date, to_date, ws_front_sheet, client_identifier):
    timestamp = date.today().strftime('%Y-%m-%d')
    start_date = from_date.strftime('%Y-%m-%d')
    end_date = to_date.strftime('%Y-%m-%d')

    header_info = [
        ("INDLUE (PTY) LTD - Cell no", client_identifier),
        ("Monthly Insurance Remittance", ""),
        ("Remittance No.", ""),
        ("Date:", timestamp),
        ("", ""),
        ("Bill To:", ""),
        ("", "43 Ingersol Street"),
        ("", "Lynwood Glen"),
        ("", "Pretoria"),
        ("Address:", "81"),
        ("", ""),
        ("Phone:", "(011) 669-1070"),
        ("E-mail:", ""),
        ("", ""),
        ("Accounting period:  {} to {}".format(start_date, end_date), ""),
    ]
    for row_num, (left, right) in enumerate(header_info, 1):
        ws_front_sheet.cell(row=row_num, column=1, value=left)
        ws_front_sheet.cell(row=row_num, column=2, value=right)

    # Write the table header
    table_header = [
        "Entity", "Risk Premium", "Guardrisk Fee", "Commission", "Binder Fee", "Nett"
    ]
    ws_front_sheet.append(table_header)
    for col in range(1, len(table_header) + 1):
        ws_front_sheet.cell(row=len(header_info) + 1, column=col).font = Font(bold=True)

    for entity, sums in reports.items():
        ws_front_sheet.append([entity] + sums)

    for row in totals:
        ws_front_sheet.append(row)


def populate_policies(policy_payments, from_date, to_date, entity, timestamp):
    flattened_data = []
    for data in policy_payments:
        policy_number = data["policy_number_annotated"]
        premium = float(data["premium_annotated"])
        premium_amount_paid = float(data["premium_paid"])
        vat_amount = calculate_vat_amount(premium_amount_paid)
        premium_less_vat = calculate_amount_excluding_vat(premium_amount_paid, vat_amount)
        guardrisk_amount = calculate_guard_risk_admin_amount(premium_amount_paid)
        commission = calculate_insurer_commission_amount(premium_amount_paid)
        binder_fee = calculate_binder_fee_amount(premium_amount_paid)
        nett_amount = calculate_nett_amount(premium_amount_paid, guardrisk_amount, commission, binder_fee)
        commencement_date = data["commencement_date_annotated"]
        business_unit = data["division"]
        policy_term = data["policy_term_annotated"]
        first_name = data["first_name"]
        last_name = data["last_name"]
        initials = f"{first_name[0]}{last_name[0]}"
        flattened_item = {
            "timestamp": timestamp,
            "period_start": from_date,
            "period_end": to_date,
            "administrator_identifier": "Indlu",
            "insurer": "Guardrisk",
            "client_identifier": get_client_identifier(entity),
            "division": data["division"],
            "risk_identifier": data["risk_identifier_annotated"] or generate_risk_identifier(business_unit,
                                                                                             policy_term),
            "sub_scheme": data["scheme_sub_annotated"],
            "policy_number": policy_number,
            "commencement_date": data["commencement_date_annotated"],
            "expiry_date": data["expiry_date_annotated"],
            "new_business_indicator": is_new_policy(commencement_date, from_date, to_date),
            "policy_term": data["policy_term_annotated"],
            "premium": premium,
            "premium_paid": premium_amount_paid,
            "commission_amount": commission,
            "binder_fee": binder_fee,
            "net_premium_paid_to_gr": nett_amount,
            "admin_fee": guardrisk_amount,
            "premium_frequency": data["premium_frequency_annotated"],
            "death_indicator": "Y",
            "ptd_indicator": "Y",
            "ttd_indicator": "Y",
            "retrenchment_indicator": "Y",
            "dread_disease_indicator": "N",
            "identity_theft_indicator": "N",
            "accidental_death_indicator": "N",
            "sum_insured": data["sum_insured_annotated"],
            "current_outstanding_balance": data["current_outstanding_balance_annotated"],
            "installment_amount": data["installment_amount_annotated"],
            "last_name": last_name,
            "first_name": first_name,
            "initials": initials,
            "primary_id_number": data["primary_id_number"],
            "gender": data["gender"],
            "date_of_birth": data["date_of_birth"],
            "date_of_death": data["date_of_death"],
            "address_street": data["address_street"],
            "address_town": data["address_town"],
            "address_province": data["address_province"],
            "postal_code": data["postal_code"],
            "phone_number": data["phone_number"]
        }
        flattened_data.append(flattened_item)
    return flattened_data


def get_client_identifier(entity):
    return "75" if entity == "Indlu" else "143"


def generate_risk_identifier(business_unit, policy_term):
    return f"{business_unit}-{policy_term}M"


def extract_dictionary(data):
    return list(data.values())


def convert_dictionary_to_list(dictionaries):
    returned_data = []
    for data in dictionaries:
        data = extract_dictionary(data)
        returned_data.append(data)
    return returned_data
