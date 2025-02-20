import calendar
from collections import defaultdict
from datetime import date
from io import BytesIO

import openpyxl
import pandas as pd
from django.db.models import Q, Count, Case, When, Value, CharField
from django.db.models import Sum
from django.db.models.functions import TruncMonth

from claims.models import Claim
from clients.models import ClientDetails
from integrations.guardrisk.data.premiums import calculate_binder_fee_amount, calculate_insurer_commission_amount
from policies.models import Policy, PremiumPayment
from reports.utils import get_client_identifier, convert_dictionary_to_list


def fetch_active_policies_as_at_date(entity, given_date):
    active_policies = Policy.objects.filter(
        Q(entity=entity),
        Q(commencement_date__lte=given_date),  # Start date is before or on the given date
        (Q(closed_date__gte=given_date) | Q(closed_date__isnull=True))  # Policy is not closed by the given date
    )
    return active_policies


def fetch_new_policies_between(entity, start_date, end_date):
    new_policies = Policy.objects.filter(
        Q(entity=entity),
        Q(commencement_date__range=(start_date, end_date))
    )
    return new_policies


def fetch_lapsed_policies_between(entity, start_date, end_date):
    lapsed_policies = Policy.objects.filter(
        Q(entity=entity),
        Q(closed_date__range=(start_date, end_date))
    )
    return lapsed_policies


def fetch_quarterly_bordraux_summary(from_date, to_date, entity):
    active_policies_period_start = fetch_active_policies_as_at_date(entity, given_date=from_date)
    active_policies_period_end = fetch_active_policies_as_at_date(entity, given_date=to_date)
    new_policies = fetch_new_policies_between(entity, from_date, to_date)
    categorized_policies = get_categorized_policy_counts(new_policies)

    categorized_policies_list = [
        {"category": entry["category"], "count": entry["count"]}
        for entry in categorized_policies
    ]
    lapsed_policies = fetch_lapsed_policies_between(entity, from_date, to_date)
    summaries = sum_and_get_summaries(new_policies, active_policies_period_start, active_policies_period_end,
                                      lapsed_policies)
    active_start_premium = summaries['active_start_premium']
    active_start_insured = summaries['active_start_insured']
    active_end_premium = summaries['active_end_premium']
    active_end_insured = summaries['active_end_insured']
    new_premium = summaries['new_premium']
    new_insured = summaries['new_insured']
    lapsed_premium = summaries['lapsed_premium']
    lapsed_insured = summaries['lapsed_insured']
    return {
        "active_policies_initially_count": len(active_policies_period_start),
        "active_policies_initially_premium": active_start_premium,
        "active_policies_initially_annual_premium": active_start_premium * 12,
        "active_policies_initially_sum_insured": active_start_insured,

        "active_policies_end_count": len(active_policies_period_end),
        "active_policies_end_premium": active_end_premium,
        "active_policies_end_annual_premium": active_end_premium * 12,
        "active_policies_end_sum_insured": active_end_insured,

        "new_policies_count": len(new_policies),
        "new_policies_premium": new_premium,
        "new_policies_annual_premium": new_premium * 12,
        "new_policies_sum_insured": new_insured,

        "lapsed_policies_count": len(lapsed_policies),
        "lapsed_policies_premium": lapsed_premium,
        "lapsed_policies_annual_premium": lapsed_premium * 12,
        "lapsed_policies_sum_insured": lapsed_insured,
        "categorized_policies": categorized_policies_list
    }


def get_categorized_policy_counts(policies):
    categorized_policies = policies.annotate(
        category=Case(
            When(policy_details__score__gte=673, policy_details__score__lte=900, then=Value('CatA')),
            When(policy_details__score__gte=652, policy_details__score__lte=672, then=Value('CatB')),
            When(policy_details__score__gte=1, policy_details__score__lte=651, then=Value('CatC')),
            default=Value('Other'),
            output_field=CharField(),
        )
    ).values('category').annotate(count=Count('id')).order_by('category')

    return categorized_policies


def generate_quarterly_excel_report(from_date, to_date, entity):
    active_policies_period_start = fetch_active_policies_as_at_date(entity, given_date=from_date)
    active_policies_period_end = fetch_active_policies_as_at_date(entity, given_date=to_date)
    new_policies = fetch_new_policies_between(entity, from_date, to_date)
    category_counts = get_categorized_policy_counts(new_policies)
    lapsed_policies = fetch_lapsed_policies_between(entity, from_date, to_date)
    wb = openpyxl.Workbook()
    ws_front_sheet = wb.active
    ws_front_sheet.title = "Quarter Summary"

    ws_new_policies_sheet = wb.create_sheet(title="New Policies")
    ws_lapsed_policies_sheet = wb.create_sheet(title="Lapsed Policies")
    ws_active_policies_beginning_sheet = wb.create_sheet(title="Active Policies beginning")
    ws_active_policies_at_end_sheet = wb.create_sheet(title="Active Policies end")
    timestamp = date.today().strftime("%Y%m%d")
    client_identifier = get_client_identifier(entity)
    new_policies_populated = populate_policies(new_policies, entity, timestamp)
    active_policies_beginning_quarter_populated = populate_policies(active_policies_period_start, client_identifier,
                                                                    timestamp)
    active_policies_end_quarter_populated = populate_policies(active_policies_period_end, client_identifier, timestamp)
    lapsed_policies_populated = populate_policies(lapsed_policies, client_identifier, timestamp)
    generate_summary_sheet(
        ws_front_sheet,
        new_policies,
        active_policies_period_start,
        active_policies_period_end,
        lapsed_policies,
        category_counts,
    )
    header = ["Time Stamp", "Administrator Identifier", "Insurer Name",
              "Client Identifier", "Division", "Policy Type", "Sub Scheme Name", "Policy Number",
              "Commencement Date", "Expiry Date", "Policy Term", "Premium", "Sum Insured"
              ]
    populate_data_sheet(new_policies_populated, header, ws_new_policies_sheet)
    populate_data_sheet(lapsed_policies_populated, header, ws_lapsed_policies_sheet)
    populate_data_sheet(active_policies_beginning_quarter_populated, header, ws_active_policies_beginning_sheet)
    populate_data_sheet(active_policies_end_quarter_populated, header, ws_active_policies_at_end_sheet)

    return write_work_book_bytes(workbook=wb)


def populate_data_sheet(data, header, worksheet):
    worksheet.append(header)
    data = convert_dictionary_to_list(data)
    for row in data:
        worksheet.append(row)


def generate_summary_sheet(summary_sheet, new_policies, active_policies_period_start, active_policies_period_end,
                           lapsed_policies, categorized_policies):
    summaries = sum_and_get_summaries(new_policies, active_policies_period_start, active_policies_period_end,
                                      lapsed_policies)
    active_start_premium = summaries['active_start_premium']
    active_start_insured = summaries['active_start_insured']
    active_end_premium = summaries['active_end_premium']
    active_end_insured = summaries['active_end_insured']
    new_premium = summaries['new_premium']
    new_insured = summaries['new_insured']
    lapsed_premium = summaries['lapsed_premium']
    lapsed_insured = summaries['lapsed_insured']

    new_policies_data = ["New Policies", len(new_policies), f"{new_premium:,.2f}", f"{new_premium * 12:,.2f}",
                         f"{new_insured:,.2f}"]
    active_policies_at_start = ["Active Policies Start", len(active_policies_period_start),
                                f"{active_start_premium:,.2f}", f"{active_start_premium * 12:,.2f}",
                                f"{active_start_insured:,.2f}"]
    active_policies_at_end = ["Active Policies End", len(active_policies_period_end), f"{active_end_premium:,.2f}",
                              f"{active_end_premium * 12:,.2f}", f"{active_end_insured:,.2f}"]
    lapsed_policies_data = ["Lapsed Policies", len(lapsed_policies), f"{lapsed_premium:,.2f}",
                            f"{lapsed_premium * 12:,.2f}",
                            f"{lapsed_insured:,.2f}"]
    categorized_policies_list = [
        [entry["category"], entry["count"]]
        for entry in categorized_policies
    ]
    for policy in categorized_policies_list:
        print(f'categorized in summary sheet policy: {policy}')
    headers = ["Policy Type", "Count", "Premium", "Annual Premium", "Sum Insured"]
    summary_sheet.append(headers)
    summary_sheet.append(new_policies_data)
    summary_sheet.append(lapsed_policies_data)
    summary_sheet.append(active_policies_at_start)
    summary_sheet.append(active_policies_at_end)
    for policy in categorized_policies_list:
        summary_sheet.append(policy)


def generate_policy_summary_sheet(summary_sheet, policies):
    summaries = summarize_policies(policies)
    premium = summaries['premium']
    sum_insured = summaries['sum_insured']
    count = summaries['policy_count']

    new_policies_data = ["Policies", count, premium, sum_insured]

    headers = ["Count", "Premium", "Sum Insured"]
    summary_sheet.append(headers)
    summary_sheet.append(new_policies_data)


def sum_and_get_summaries(
        new_policies,
        active_policies_period_start,
        active_policies_period_end,
        lapsed_policies):
    active_start_premium = sum(policy.premium for policy in active_policies_period_start)
    active_start_insured = sum(policy.sum_insured for policy in active_policies_period_start)
    active_end_premium = sum(policy.premium for policy in active_policies_period_end)
    active_end_insured = sum(policy.sum_insured for policy in active_policies_period_end)
    new_premium = sum(policy.premium for policy in new_policies)
    new_insured = sum(policy.sum_insured for policy in new_policies)
    lapsed_premium = sum(policy.premium for policy in lapsed_policies)
    lapsed_insured = sum(policy.sum_insured for policy in lapsed_policies)
    return {
        "active_start_premium": active_start_premium,
        "active_start_insured": active_start_insured,
        "active_end_premium": active_end_premium,
        "active_end_insured": active_end_insured,
        "new_premium": new_premium,
        "new_insured": new_insured,
        "lapsed_premium": lapsed_premium,
        "lapsed_insured": lapsed_insured
    }


def populate_policies(policies, entity, timestamp):
    flattened_data = []
    for data in policies:
        policy_number = data.policy_number
        premium = float(data.premium)
        sum_assured = float(data.sum_insured)
        commencement_date = data.commencement_date
        business_unit = data.business_unit
        policy_term = data.policy_term
        policy_type = data.policy_type.name
        flattened_item = {
            "timestamp": timestamp,
            "administrator_identifier": data.entity,
            "insurer": "Guardrisk",
            "client_identifier": entity,
            "division": business_unit,
            "policy_type": policy_type,
            "sub_scheme": data.sub_scheme,
            "policy_number": policy_number,
            "commencement_date": commencement_date,
            "expiry_date": data.expiry_date,
            "policy_term": policy_term,
            "premium": premium,
            "sum_insured": sum_assured
        }
        flattened_data.append(flattened_item)
    return flattened_data


def flatten_policies(policies, timestamp):
    flattened_data = []
    for data in policies:
        policy_number = data.policy_number
        premium = float(data.premium)
        sum_assured = float(data.sum_insured)
        commencement_date = data.commencement_date
        business_unit = data.business_unit
        policy_term = data.policy_term
        policy_type = data.policy_type.name
        flattened_item = {
            "timestamp": timestamp,
            "administrator_identifier": data.entity,
            "insurer": "Guardrisk",
            "division": business_unit,
            "policy_type": policy_type,
            "product_name": data.product_name,
            "sub_scheme": data.sub_scheme,
            "policy_number": policy_number,
            "commencement_date": commencement_date,
            "expiry_date": data.expiry_date,
            "closed_date": data.closed_date,
            "policy_status": data.policy_status,
            "policy_term": policy_term,
            "premium": premium,
            "sum_insured": sum_assured
        }
        flattened_data.append(flattened_item)
    return flattened_data


def generate_policy_report(policies):
    aggregates = defaultdict(lambda: [0, 0, 0])
    for policy in policies:
        division = policy['division']
        print(f'division: {division}')
        try:
            risk_premium = float(policy['premium'])
        except TypeError as e:
            risk_premium = 0
        try:
            annual_risk_premium = float(policy['annual_sum_insured'])
        except TypeError as e:
            annual_risk_premium = 0
        try:
            total_sum_assured = float(policy['sum_insured'])
        except TypeError as e:
            total_sum_assured = 0
        else:
            aggregates[division][0] += risk_premium
            aggregates[division][1] += annual_risk_premium
            aggregates[division][2] += total_sum_assured
    totals = {
        "division": "",
        "total_risk_premium": 0,
        "total_annual_risk_premium": 0,
        "total_sum_assured": 0
    }
    for division, tot in aggregates.items():
        totals["division"] = division
        totals["total_risk_premium"] += tot[0]
        totals["total_annual_risk_premium"] += tot[1]
        totals["total_sum_assured"] += tot[2]
    return aggregates, totals


def fetch_policies(policy_type, from_date, to_date, query):
    policies = Policy.objects.all()
    if query:
        policies = policies.filter(
            Q(client__first_name__icontains=query)
            | Q(client__last_name__icontains=query)
            | Q(client__middle_name__icontains=query)
            | Q(client__external_id__icontains=query)
            | Q(client__email__icontains=query)
            | Q(client__phone_number__icontains=query)
            | Q(client__primary_id_number__icontains=query)
            | Q(insurer__name__icontains=query)
            | Q(policy_number__icontains=query)
            | Q(external_reference__icontains=query)
        )

    if policy_type is not None:
        policies = policies.filter(policy_type__id=policy_type)

    if from_date:
        policies = policies.filter(commencement_date__gte=from_date)

    if to_date:
        policies = policies.filter(commencement_date__lte=to_date)
    return policies


def summarize_policies(policies):
    df = pd.DataFrame(list(policies.values('premium', 'sum_insured')))
    if df.empty:
        return {
            'premium': 0,
            'sum_insured': 0,
            'policy_count': 0
        }
    summary = df.agg({
        'premium': 'sum',
        'sum_insured': 'sum'
    }).to_dict()
    summary['policy_count'] = len(df)
    return summary


def generate_policies_excel_report(policy_type, from_date, to_date, query):
    policies = fetch_policies(policy_type, from_date, to_date, query)
    wb = openpyxl.Workbook()
    ws_front_sheet = wb.active
    ws_front_sheet.title = "Policies Summary"

    ws_policies_sheet = wb.create_sheet(title="Policies")
    timestamp = date.today().strftime("%Y%m%d")
    policies_flattened = flatten_policies(policies, timestamp)
    generate_policy_summary_sheet(
        ws_front_sheet,
        policies
    )
    header = ["Time Stamp", "Administrator Identifier", "Insurer Name",
              "Division", "Policy Type", "Product Name", "Sub Scheme Name", "Policy Number",
              "Commencement Date", "Expiry Date", "Closed Date", "Status", "Policy Term", "Premium", "Sum Insured"
              ]
    populate_data_sheet(policies_flattened, header, ws_policies_sheet)
    return write_work_book_bytes(wb)


def write_work_book_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_claims_excel_report(claim_type, from_date, to_date, query):
    claims = fetch_claims(claim_type, from_date, to_date, query)
    wb = openpyxl.Workbook()
    ws_front_sheet = wb.active
    ws_front_sheet.title = "Claims Summary"

    ws_claims_sheet = wb.create_sheet(title="Claims")
    timestamp = date.today().strftime("%Y%m%d")
    claims_flattened = flatten_claims(claims, timestamp)
    generate_claim_summary_sheet(
        ws_front_sheet,
        claims
    )
    header = ["Time Stamp", "Policy Number", "Claim Type",
              "Status", "Claimed Amount", "Sum Insured", "Division", "Policy Type", "Product Name",
              "Commencement Date", "Policy Term", "Submitted Date"
              ]
    populate_data_sheet(claims_flattened, header, ws_claims_sheet)
    return write_work_book_bytes(wb)


def fetch_claims(claim_type, from_date, to_date, query):
    print('fetching...')
    claims = Claim.objects.all()
    if query:
        claims = claims.filter(
            Q(claimant_name__icontains=query)
            | Q(claimant_surname__icontains=query)
            | Q(claimant_id_number__icontains=query)
            | Q(claimant_email__icontains=query)
            | Q(claimant_phone__icontains=query)
            | Q(policy__insurer__name__icontains=query)
            | Q(claim_status__icontains=query)
        )

    if claim_type is not None:
        claims = claims.filter(claim_type__id=claim_type)
    if from_date:
        claims = claims.filter(created__gte=from_date)

    if to_date:
        claims = claims.filter(created__lte=to_date)
    return claims


def flatten_claims(claims, timestamp):
    flattened_data = []
    for data in claims:
        policy_number = data.policy.policy_number
        claimed_amount = float(data.claim_amount)
        sum_assured = float(data.policy.sum_insured)
        commencement_date = data.policy.commencement_date
        business_unit = data.policy.business_unit
        policy_term = data.policy.policy_term
        policy_type = data.policy.policy_type.name
        claim_type = data.claim_type.name
        claim_status = data.claim_status
        submitted_date = data.submitted_date
        flattened_item = {
            "timestamp": timestamp,
            "policy_number": policy_number,
            "claim_type": claim_type,
            "claim_status": claim_status,
            "claimed_amount": claimed_amount,
            "sum_insured": sum_assured,
            "division": business_unit,
            "policy_type": policy_type,
            "product_name": data.policy.product_name,
            "commencement_date": commencement_date,
            "policy_term": policy_term,
            "submitted_date": submitted_date
        }
        flattened_data.append(flattened_item)
    return flattened_data


def generate_claim_summary_sheet(summary_sheet, claims):
    summaries = summarize_claims(claims)
    total_claim_amount = summaries['total_claim_amount']
    count = summaries['claim_count']
    breakdown_by_status = summaries['breakdown_by_status']

    claims_data = ["Claims", count, total_claim_amount]

    headers = ["Count", "Total Claimed Amount"]
    summary_sheet.append(headers)

    breakdown_by_status = [
        [entry['claim_status'], entry['claim_status_count'], entry['total_status_claim_amount'] or 0.0]
        for entry in breakdown_by_status
    ]
    summary_sheet.append(breakdown_by_status)
    summary_sheet.append(claims_data)


def summarize_claims(claims):
    total_summary = claims.aggregate(
        total_claim_amount=Sum('claim_amount'),
        total_claim_count=Count('id')
    )
    status_breakdown = claims.values('claim_status').annotate(
        claim_status_count=Count('id'),
        total_status_claim_amount=Sum('claim_amount')
    )
    summary = {
        "claim_count": total_summary['total_claim_count'],
        "total_claim_amount": total_summary['total_claim_amount'] or 0.0,
        "breakdown_by_status": list(status_breakdown)
    }
    return summary


def summarize_clients(clients):
    df = pd.DataFrame(list(clients.values()))
    if df.empty:
        return {
            'client_count': 0
        }
    return {'client_count': len(df)}


def generate_clients_excel_report(clients):
    wb = openpyxl.Workbook()
    ws_front_sheet = wb.active
    ws_front_sheet.title = "Clients Summary"

    ws_clients_sheet = wb.create_sheet(title="Clients")
    timestamp = date.today().strftime("%Y%m%d")
    clients_flattened = flatten_clients(clients, timestamp)
    generate_clients_summary_sheet(
        ws_front_sheet,
        clients
    )
    header = ["Time Stamp", "Title", "Id number",
              "First Name", "Middle Name", "Surname", "Entity Type", "Gender", "Email", "Contact",
              "Marital Status", "D.O.B", "D.O.D", "Address"
              ]
    populate_data_sheet(clients_flattened, header, ws_clients_sheet)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def flatten_clients(clients, timestamp):
    flattened_data = []
    for data in clients:
        id_number = data.primary_id_number
        flattened_item = {
            "timestamp": timestamp,
            "title": data.title,
            "id_number": id_number,
            "first_name": data.first_name,
            "middle_name": data.middle_name,
            "last_name": data.last_name,
            "entity_type": data.entity_type,
            "gender": data.gender,
            "email": data.email,
            "phone_number": data.phone_number,
            "marital_status": data.marital_status,
            "date_of_birth": data.date_of_birth,
            "date_of_death": data.date_of_death,
            "address": (
                f"{data.address_street or ''} {data.address_suburb or ''} "
                f"{data.address_town or ''}, {data.address_province or ''}"
            ).strip()
        }
        flattened_data.append(flattened_item)
    return flattened_data


def generate_clients_summary_sheet(summary_sheet, clients):
    summaries = summarize_clients(clients)
    count = summaries['client_count']

    clients_data = ["Clients", count]

    headers = ["Count"]
    summary_sheet.append(headers)
    summary_sheet.append(clients_data)


def fetch_dashboard_stats(from_date, to_date, policy_type):
    print(f'policy_type: {policy_type}')
    # active_policies_period_start = fetch_active_policies_as_at_date(entity, given_date=from_date)
    # active_policies_period_end = fetch_active_policies_as_at_date(entity, given_date=to_date)
    new_policies = fetch_all_new_policies(from_date, to_date, policy_type)
    # lapsed_policies = fetch_lapsed_policies_between(entity, from_date, to_date)
    clients_by_gender = fetch_clients_summarized_by_gender()

    df = pd.DataFrame(list(new_policies.values('premium', 'sum_insured')))
    if df.empty:
        return {
            'premium': 0,
            'sum_insured': 0,
            'policy_count': 0,
            'total_commission': 0,
            'binder_fee': 0,
            'income': []
        }
    df['commission'] = df['premium'].astype(float).apply(calculate_insurer_commission_amount)
    df['binder_fee'] = df['premium'].astype(float).apply(calculate_binder_fee_amount)
    summary = df.agg({
        'premium': 'sum',
        'sum_insured': 'sum',
        'commission': 'sum',
        'binder_fee': 'sum'
    }).to_dict()
    summary['policy_count'] = len(df)
    expected_premiums = calculate_total_premium_for_active_policies(start_date=from_date, end_date=to_date)
    received_premiums = get_total_payments_by_month(from_date, to_date)
    combined = combine_premium_data(expected_premiums, received_premiums)
    summary['income'] = combined
    summary['clients_by_gender'] = clients_by_gender
    return summary


def fetch_all_new_policies(from_date, to_date, policy_type):
    new_policies = Policy.objects.filter(
        Q(policy_type__id=policy_type),
        Q(commencement_date__range=(from_date, to_date))
    )
    return new_policies


def fetch_clients_summarized_by_gender():
    gender_summary = (
        ClientDetails.objects
        .values('gender')
        .annotate(client_count=Count('id'))
        .order_by('gender')
    )

    return gender_summary


def calculate_total_premium_for_active_policies(start_date: date, end_date: date):
    """
    Calculate the total premium for active policies by month within a given period.
    Only pick the policies that are active during the specified period.
    """
    # Step 1: Filter policies that are active in the given period
    active_policies = active_policies_during_period(start_date, end_date)

    total_premium_by_month = (
        active_policies
        .annotate(month=TruncMonth('commencement_date'))
        .values('month')
        .annotate(premium=Sum('premium'))
        .order_by('month')
    )

    return total_premium_by_month


def active_policies_during_period(start_date, end_date):
    return Policy.objects.filter(
        Q(commencement_date__lte=end_date),
        Q(closed_date__isnull=True) | Q(closed_date__gte=start_date)
    )


def get_total_payments_by_month(start_date: date, end_date: date):
    payments_by_month = (
        PremiumPayment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date
        )
        .annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(amount_paid=Sum('amount'))
        .order_by('month')
    )

    return payments_by_month


def combine_premium_data(expected_premiums, received_premiums):
    combined_data = {entry['month'].strftime('%b'): {'expected': entry['premium']} for entry in expected_premiums}

    for entry in received_premiums:
        month_key = entry['month'].strftime('%b')
        if month_key in combined_data:
            combined_data[month_key]['received'] = entry['amount_paid']
        else:
            combined_data[month_key] = {'expected': 0, 'received': entry['amount_paid']}

    result = [
        {
            'name': month,
            'received': combined_data[month].get('received', 0),
            'expected': combined_data[month].get('expected', 0),
        }
        for month in calendar.month_abbr[1:]
    ]
    return result
